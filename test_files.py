import os
from zipfile import ZipFile
from PyPDF2 import PdfReader
from openpyxl import load_workbook

import zipfile
from constants import CURRENT_DIR
from io import BytesIO

def test_text():

    if not os.path.exists("resources"):
        os.mkdir("resources")

    #создаем архив
    tmp_dir = os.path.join(CURRENT_DIR, "tmp")
    with zipfile.ZipFile("resources/test.zip", 'w', \
                         compression=zipfile.ZIP_STORED) as zf:
        files = os.listdir(tmp_dir)
        for file in files:
            add_file = os.path.join("tmp", file)
            zf.write(add_file)


    zip=ZipFile("resources/test.zip","r")

    #проверка текстового файла
    textdata = zip.read('tmp/text_test.txt')
    assert "You are free to use this image" in textdata.decode(), "txt-file не прошел валидацию"

    #проверка pdf
    pdf_file = PdfReader(BytesIO(zip.read("tmp/pdf_test.pdf")))
    number_of_pdf_pages = len(pdf_file.pages)
    page = pdf_file.pages[0]
    pdftext = page.extract_text()
    assert number_of_pdf_pages == 39
    assert "Права и обязанности сторон" in pdftext, "PDF-file не прошел валидацию"

    #проверка xlsx
    workbook = load_workbook(BytesIO(zip.read("tmp/xls_test.xlsx")))
    sheet = workbook.active
    assert "Этапы тестирования" == sheet.cell(row=5, column=1).value, "xlsx-file не прошел валидацию"

